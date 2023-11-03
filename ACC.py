#######################################################################
# Copyright (c) 2018 Cisco Systems, Inc. - All Rights Reserved
# Unauthorized copying of this file,
# via any medium is strictly prohibited
# Proprietary and confidential
#######################################################################

from openpyxl import Workbook
import csv
import requests
import json 
import datetime
import getpass
import urllib3
from openpyxl import Workbook, load_workbook
import csv
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL:@SECLEVEL=1'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#APIC Login Function 
def apic_login(apic: str, username: str, password: str) -> dict:
    apic_cookie = {}
    credentials = {'aaaUser': {'attributes': {'name': username, 'pwd': password }}}
    json_credentials = json.dumps(credentials)
    base_url = 'https://' + apic + '/api/aaaLogin.json'
    login_response = requests.post(base_url, data=json_credentials, verify=False)
    login_response_json = json.loads(login_response.text)
    token = login_response_json['imdata'][0]['aaaLogin']['attributes']['token']
    apic_cookie['APIC-Cookie'] = token
    return apic_cookie

#APIC API Call Function 
def apic_query(apic: str, path: str, cookie: dict) -> dict:
    base_url = 'https://' + apic + path
    get_response = requests.get(base_url, cookies=cookie, verify=False)
    return get_response

#APIC Logout Function 
def apic_logout(apic: str, cookie:dict) -> dict:
    base_url = 'https://' + apic + '/api/aaaLogout.json'
    post_response = requests.post(base_url, cookies=cookie, verify=False)
    return post_response

#Gather Credentials from CSV & run API Calls
def process_credential(apic_url, username, password, fabric_name, wb):
    #Logging in & retrieving APIC token
    apic_cookie = apic_login(apic=apic_url, username=username, password=password)
    print(f"Logging into APIC {fabric_name} \n")
    print("Gathering data... \n")

    # Check if the worksheet with the same name as the fabric exists
    if fabric_name in wb.sheetnames:
        ws = wb.create_sheet(fabric_name)

    #Retrieving Leaf Count 
    leafcount = apic_query(apic=apic_url, path='/api/node/class/topology/pod-1/topSystem.json?query-target-filter=eq(topSystem.role,\"leaf\")&rsp-subtree-include=health,count', cookie=apic_cookie)
    response_json = json.loads(leafcount.text)
    leafs = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retrieving Spine Count 
    spinecount = apic_query(apic=apic_url, path='/api/node/class/topology/pod-1/topSystem.json?query-target-filter=eq(topSystem.role,\"spine\")&rsp-subtree-include=health,count', cookie=apic_cookie)
    response_json = json.loads(spinecount.text)
    spines = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Controller Count
    contr_count = apic_query(apic=apic_url, path='/api/node/class/topSystem.json?rsp-subtree-include=count&query-target-filter=eq(topSystem.role,"controller")', cookie=apic_cookie)
    response_json = json.loads(contr_count.text)
    controllers = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Tenant Count 
    ten_count = apic_query(apic=apic_url, path='/api/node/class/fvTenant.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(ten_count.text)
    tenant = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Bridge Domain Count 
    bd_count = apic_query(apic=apic_url, path='/api/node/class/fvBD.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(bd_count.text)
    bd = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving VRF Count
    vrf_count = apic_query(apic=apic_url, path='/api/node/class/fvCtx.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(vrf_count.text)
    vrf = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving EPG Count
    epg_count = apic_query(apic=apic_url, path='/api/node/class/fvAEPg.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(epg_count.text)
    epg = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Application Profiles 
    ap_count = apic_query(apic=apic_url, path='/api/node/class/fvAp.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(ap_count.text)
    ap = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Contracts 
    contrac_count = apic_query(apic=apic_url, path='/api/node/class/vzBrCP.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(contrac_count.text)
    contrac = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving L3Outs
    l3out_count = apic_query(apic=apic_url, path='/api/node/class/l3extOut.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(l3out_count.text)
    l3out = response_json["imdata"][0]["moCount"]["attributes"]["count"]

    #Retreiving Proxy Data Entries Count
    ip_count = apic_query(apic=apic_url, path='/api/node/class/fvIp.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(ip_count.text)
    Ip = response_json["imdata"][0]["moCount"]["attributes"]["count"]
    cep_count = apic_query(apic=apic_url, path='/api/node/class/fvCEp.json?rsp-subtree-include=count', cookie=apic_cookie)
    response_json = json.loads(cep_count.text)
    CEp = response_json["imdata"][0]["moCount"]["attributes"]["count"]
    dataentries = (int(Ip) + int(CEp)) # Proxy Data Entries = IP Count and CEp Count 
    dataentries = str(dataentries)

    #Retreiving Version 
    curr_release = apic_query(apic=apic_url, path='/api/node/class/topology/pod-1/node-1/firmwareCtrlrRunning.json?', cookie=apic_cookie)
    response_json = json.loads(curr_release.text)
    release = response_json["imdata"][0]["firmwareCtrlrRunning"]["attributes"]["version"]

    #Retreiving Time & Date 
    now = datetime.datetime.today() 
    date = now.strftime("%m/%d/%Y")

    # Create a new worksheet for each fabric name
    ws = wb.create_sheet()
    ws.title = fabric_name  # Set the worksheet title to the fabric name

    #Populate column headers in worksheet 
    ws['A1'].value = "Date"
    ws['B1'].value = "Release"
    ws['C1'].value = "EPGs"
    ws['D1'].value = "EPs"
    ws['E1'].value = "Tenants"
    ws['F1'].value = "VRFs"
    ws['G1'].value = "BDs"
    ws['H1'].value = "L3Outs"
    ws['I1'].value = "Contracts"
    ws['J1'].value = "Application Profiles"
    ws['K1'].value = "Leafs"
    ws['L1'].value = "Spines"
    ws['M1'].value = "Controllers"

    #Populate API data in worksheet
    ws['A2'].value = date
    ws['B2'].value = release
    ws['C2'].value = epg
    ws['D2'].value = dataentries
    ws['E2'].value = tenant
    ws['F2'].value = vrf
    ws['G2'].value = bd
    ws['H2'].value = l3out
    ws['I2'].value = contrac
    ws['J2'].value = ap
    ws['K2'].value = leafs
    ws['L2'].value = spines
    ws['M2'].value = controllers

    # Logging out of APIC
    logout_response = apic_logout(apic=apic_url, cookie=apic_cookie)
    print(f"Logging out of APIC {fabric_name} \n")

def main():
    csv_file = input("Enter the path to the CSV file with APIC credentials: ")

    #Prompt user for existing Excel File
    existing_excel = input("Do you have an existing Excel file? (yes/no): ")

    if existing_excel.lower() =="yes":
        excel_file = input("Enter the path to the existing Excel file: ")
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
    
    #Remove the default empty worksheet created by Workbook
    try:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
    except KeyError:
        pass

    #Opening CSV Credentials File
    with open(csv_file, 'r') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            apic_url = row["APIC_IP"]
            username = row["Username"]
            password = row["Password"]
            fabric_name = row["Fabric_Name"]
            process_credential(apic_url, username, password, fabric_name, wb)

    #Save the workbook back to the Excel file
    if existing_excel.lower() == "yes":
        wb.save(excel_file)
    else:
        excel_name = input("Name your Excel file: ") + ".xlsx"
        wb.save(excel_name)

    # Option to restart code
    restart = input("Do you wish to create another file? yes/no - \n ")
    if restart.lower() == "yes":
        main()
    else:
        exit()

if __name__ == "__main__":
    main()
