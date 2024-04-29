#######################################################################
# Copyright (c) 2024 Cisco Systems, Inc. - All Rights Reserved
# Unauthorized copying of this file,
# via any medium is strictly prohibited
# Proprietary and confidential
#######################################################################
#Name will be AIC (APIC Information Collector)
import csv
import json 
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
import pandas as pd
import requests
import urllib3

# Disable SSL warnings
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL:@SECLEVEL=1'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Read CSV Function
def read_csv(csv_file):
    data = []
    with open(csv_file, "r") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            data.append(row)
    return data

# Prepare Dictionaries from CSV
def prepare_dicts_from_csv(csv_file):
    csv_data = read_csv(csv_file)
    return [
        {
            "apic_ip": row["APIC_IP"],
            "username": row["Username"],
            "password": row["Password"],
            "fabric_name": row["Fabric_Name"]
        } 
        for row in csv_data
    ]

# APIC Login Function 
def apic_login(apic_url: str, username: str, password: str) -> dict:
    credentials = {'aaaUser': {'attributes': {'name': username, 'pwd': password }}}
    json_credentials = json.dumps(credentials)
    base_url = f'https://{apic_url}/api/aaaLogin.json'
    login_response = requests.post(base_url, data=json_credentials, verify=False)
    login_response.raise_for_status()  # Raise an exception if status is not 2xx
    token = login_response.json()['imdata'][0]['aaaLogin']['attributes']['token']
    return {'APIC-Cookie': token}

# APIC API Call Function 
def apic_query(apic_url: str, path: str, cookie: dict) -> dict:
    base_url = f'https://{apic_url}{path}'
    get_response = requests.get(base_url, cookies=cookie, verify=False)
    get_response.raise_for_status()  # Raise an exception if status is not 2xx
    return get_response.json()

# APIC Logout Function 
def apic_logout(apic_url: str, cookie: dict) -> None:
    base_url = f'https://{apic_url}/api/aaaLogout.json'
    requests.post(base_url, cookies=cookie, verify=False)

# Process Credential Function
def process_credential(apic_url, username, password, fabric_name, wb):
    ws = None  # Initialize ws 
    try:
        # Logging in & retrieving APIC token
        apic_cookie = apic_login(apic_url, username, password)
        print(f"Logging into APIC {fabric_name}...")
        print("Gathering data...")

        # Retrieving counts from APIC
        counts = {
            "EPGs": "/api/node/class/fvAEPg.json?rsp-subtree-include=count",
            "EPs": "/api/node/class/fvIp.json?rsp-subtree-include=count",
            "Tenants": "/api/node/class/fvTenant.json?rsp-subtree-include=count",
            "VRFs": "/api/node/class/fvCtx.json?rsp-subtree-include=count",
            "BDs": "/api/node/class/fvBD.json?rsp-subtree-include=count",
            "L3Outs": "/api/node/class/l3extOut.json?rsp-subtree-include=count",
            "Contracts": "/api/node/class/vzBrCP.json?rsp-subtree-include=count",
            "Application Profiles": "/api/node/class/fvAp.json?rsp-subtree-include=count",
            "Leafs": "/api/node/class/topology/pod-1/topSystem.json?query-target-filter=eq(topSystem.role,\"leaf\")&rsp-subtree-include=health,count",
            "Spines": "/api/node/class/topology/pod-1/topSystem.json?query-target-filter=eq(topSystem.role,\"spine\")&rsp-subtree-include=health,count",
            "Controllers": "/api/node/class/topSystem.json?rsp-subtree-include=count&query-target-filter=eq(topSystem.role,\"controller\")"
        }
        counts_data = {key: apic_query(apic_url, path, apic_cookie)['imdata'][0]['moCount']['attributes']['count'] for key, path in counts.items()}

        # Retrieve Release Version
        release_path = "/api/node/class/topology/pod-1/node-1/firmwareCtrlrRunning.json?"
        release = apic_query(apic_url, release_path, apic_cookie)['imdata'][0]['firmwareCtrlrRunning']['attributes']['version']

        # Retrieve current date
        date = datetime.datetime.today().strftime("%m/%d/%Y")

        ws = wb[fabric_name] if fabric_name in wb.sheetnames else None

        if ws is None:
            # Create a new worksheet for each fabric name
            ws = wb.create_sheet()
            ws.title = fabric_name
            # Populate column headers in worksheet
            ws.append(["Date", "Release"] + list(counts_data.keys()))

        # Populate API data in worksheet
        ws.append([date, release] + list(counts_data.values()))
        print(f"Data collected and saved to worksheet for APIC {fabric_name}.")

    except requests.exceptions.RequestException as e:
        print(f"Could not log into APIC {fabric_name}: {e}")

    finally:
        # Logging out of APIC if logged in
        if 'apic_cookie' in locals():
            apic_logout(apic_url, apic_cookie)
            print(f"Logged out of APIC {fabric_name}..\n")


def create_barchart(excel_file_path, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(excel_file_path)
    sheet = wb[sheet_name]

    # Assuming data is in the first column for labels and the second for values
    # Adjust the range as necessary based on your Excel file's structure
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)

    # Create the bar chart
    chart = BarChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "Your Chart Title"

    # Place the chart on the sheet
    sheet.add_chart(chart, "E2")  # Adjust the location as needed

    # Save the workbook
    wb.save(excel_file_path)
    
    
def main():
    try:
        csv_file = input("Enter the path to the CSV file with APIC credentials: ")
        credentials = prepare_dicts_from_csv(csv_file)

        # Prompt user for existing Excel File
        existing_excel = input("Do you have an existing Excel file? (yes/no): ")

        if existing_excel.lower() == "yes":
            excel_file = input("Enter the path to the existing Excel file: ")
            # Loads Existing Workbook
            wb = load_workbook(excel_file)
        else:
            # Create a new workbook
            wb = Workbook()

        # Remove the default empty worksheet created by Workbook
        try:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)
        except KeyError:
            pass

        # Process credentials and save to Excel
        for cred in credentials:
            process_credential(cred['apic_ip'], cred['username'], cred['password'], cred['fabric_name'], wb)
            
        # Save the workbook back to the Excel file
        if existing_excel.lower() == "yes":
            wb.save(excel_file)
        else:
            excel_name = input("Name your Excel file: ") + ".xlsx"
            wb.save(excel_name)
            

        # Option to restart code
        restart = input("Do you wish to create another file? (yes/no): \n")
        if restart.lower() == "yes":
            main()

    except KeyboardInterrupt:
        print("Process interrupted by the user.")

if __name__ == "__main__":
    main()