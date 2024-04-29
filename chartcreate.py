import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Load the Excel file into a DataFrame
excel_file_path = input("Enter the path for your excel file: ")  # Update this with your file's path
df = pd.read_excel(excel_file_path)

# Assume 'df' is your DataFrame after any processing you need
# For demonstration, let's assume 'df' has columns 'Date', 'Value1', 'Value2'

# Drop the 'Release' column from the DataFrame
df = df.drop(columns=['Release'])

# Convert 'df' to an 'openpyxl' workbook
wb = Workbook()
ws = wb.active

# Append DataFrame rows to Excel workbook
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Create a bar chart with openpyxl
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Data Visualization"
chart.y_axis.title = 'Values'
chart.x_axis.title = 'Date'

# Assuming 'Date' is in the first column and the data to chart is in the next columns
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=len(df.columns))
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Place the chart starting from cell E2 (for example)
ws.add_chart(chart, "E2")

# Save the workbook
new_excel_file_path = 'ExcelData_Chart.xlsx'  # Update this with your desired file path and name
wb.save(new_excel_file_path)


